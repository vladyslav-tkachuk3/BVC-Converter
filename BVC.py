import re

from os import path
from os import remove
from sys import argv
from string import punctuation
from collections import defaultdict

import tkinter
import openpyxl
import xlrd

__DEBUG_BAT = True
__DEBUG_REMAP = True
__DEBUG_ALERT_FORM = True
__DEBUG_OUTPUT = True


class Loader:
    _XLSX_FORMAT = 'xlsx'
    _XLS_FORMAT = 'xls'

    def __init__(self, path):

        # CREATE EMPTY DATA
        self._load_data = defaultdict(lambda: None)
        self._load_data['is_valid_file'] = True

        # OPEN FILE
        try:
            workbook = openpyxl.load_workbook(path)
            self._sheet = workbook.active
        except openpyxl.utils.exceptions.InvalidFileException:
            try:
                workbook = xlrd.open_workbook(path)
                self._sheet = workbook.sheet_by_index(0)
            except FileNotFoundError:
                self._load_data['is_valid_file'] = False
        except (Exception, FileNotFoundError):
            self._load_data['is_valid_file'] = False

        # DEFINE FORMAT
        if self._load_data['is_valid_file']:
            if type(self._sheet) == openpyxl.worksheet.worksheet.Worksheet:
                self._load_data['format'] = Loader._XLSX_FORMAT
            elif type(self._sheet) == xlrd.sheet.Sheet:
                self._load_data['format'] = Loader._XLS_FORMAT

        # DEFINE SIZE
        if self._load_data['is_valid_file'] and (self._load_data['format'] == Loader._XLSX_FORMAT or
                                                 self._load_data['format'] == Loader._XLS_FORMAT):
            if self._load_data['format'] == Loader._XLSX_FORMAT:
                self._load_data['size'] = self._sheet.max_row
            elif self._load_data['format'] == Loader._XLS_FORMAT:
                self._load_data['size'] = self._sheet.nrows

    def get_data(self):
        return self._load_data


class RemapLoader(Loader):

    def __init__(self, path):
        super().__init__(path)

        self.__headers = ['Name Target', 'Value', 'Label', 'Values', 'Type', 'Val Width', 'Decimals', 'Command',
                          'Command Var', '', 'Name Source', 'Value', 'Label', 'Values', 'Type', 'Val Width', 'Decimals']

        if self._load_data['is_valid_file'] and (self._load_data['format'] == Loader._XLSX_FORMAT or
                                                 self._load_data['format'] == Loader._XLS_FORMAT):
            # CHECK HEADERS AND SIZE
            if self._load_data['size'] > 6:
                self._load_data['is_empty'] = False
                self._load_data['is_valid_header'] = True
                if self._load_data['format'] == Loader._XLSX_FORMAT:
                    for i, v in enumerate(self.__headers):
                        if not len(v):
                            self.__headers[i] = None
                if self._load_data['format'] == Loader._XLSX_FORMAT:
                    for _ in range(1, len(self.__headers) + 1):
                        if self._sheet.cell(6, _).value != self.__headers[_ - 1]:
                            self._load_data['is_valid_header'] = False
                            break
                elif self._load_data['format'] == Loader._XLS_FORMAT:
                    for _ in range(0, len(self.__headers)):
                        if self._sheet.cell_value(5, _) != self.__headers[_]:
                            self._load_data['is_valid_header'] = False
                            break

                # LOAD VARIABLES
                self._load_data['vars'] = []
                if self._load_data['format'] == Loader._XLSX_FORMAT:
                    for _ in range(7, self._load_data['size'] + 1):
                        cell_value = self._sheet.cell(_, 1).value
                        if cell_value is not None and cell_value != '':
                            self._load_data['vars'].append([cell_value, _])
                    for _ in range(len(self._load_data['vars']) - 1):
                        self._load_data['vars'][_].append(self._load_data['vars'][_ + 1][1] - 1)
                    self._load_data['vars'][-1].append(self._load_data['vars'][-1][1])

                elif self._load_data['format'] == Loader._XLS_FORMAT:
                    for _ in range(6, self._load_data['size']):
                        cell_value = self._sheet.cell_value(_, 0)
                        if cell_value is not None and cell_value != '':
                            self._load_data['vars'].append([cell_value, _ + 1])
                    for _ in range(len(self._load_data['vars']) - 1):
                        self._load_data['vars'][_].append(self._load_data['vars'][_ + 1][1] - 1)
                    self._load_data['vars'][-1].append(self._load_data['vars'][-1][1])

                self._load_data['sheet'] = self._sheet
            else:
                self._load_data['is_empty'] = True


class AlertFormLoader(Loader):

    def __init__(self, path):
        super().__init__(path)

        if self._load_data['is_valid_file'] and (self._load_data['format'] == Loader._XLSX_FORMAT or
                                                 self._load_data['format'] == Loader._XLS_FORMAT):
            self._load_data['is_empty'] = not bool(self._load_data['size'])

            # CHECK INDEXES IN BAT FILE
            self._load_data['index_errors'] = []

            # CHECK BRAND LIST INDEXES
            if not argv[6].isnumeric() or int(argv[6]) <= 0:
                self._load_data['index_errors'].append(('BRAND_LIST_ROW', argv[6]))
            if not argv[7].isnumeric() or int(argv[7]) <= 0:
                self._load_data['index_errors'].append(('BRAND_LIST_COL', argv[7]))

            # CHECK EFFECTS LIST INDEXES
            if not argv[8].isnumeric() or int(argv[8]) <= 0:
                self._load_data['index_errors'].append(('MARKET_EFFECTS_ROW', argv[8]))
            if not argv[9].isnumeric() or int(argv[9]) <= 0:
                self._load_data['index_errors'].append(('MARKET_EFFECTS_COL', argv[9]))

            # CHECK BARRIERS INDEXES
            if not argv[10].isnumeric() or int(argv[10]) <= 0:
                self._load_data['index_errors'].append(('BARRIERS_CON_ROW', argv[10]))
            if not argv[11].isnumeric() or int(argv[11]) <= 0:
                self._load_data['index_errors'].append(('BARRIERS_CON_COL', argv[11]))

            # CHECK IMAGERY INDEXES
            if not argv[12].isnumeric() or int(argv[12]) <= 0:
                self._load_data['index_errors'].append(('IMAGERY_ROW', argv[12]))
            if not argv[13].isnumeric() or int(argv[13]) <= 0:
                self._load_data['index_errors'].append(('IMAGERY_COL', argv[13]))

            # CHECK DATA INFO INDEXES
            if not argv[14].isnumeric() or int(argv[14]) <= 0:
                self._load_data['index_errors'].append(('DATA_INFORMATION_ROW', argv[14]))
            if not argv[15].isnumeric() or int(argv[15]) <= 0:
                self._load_data['index_errors'].append(('DATA_INFORMATION_COL', argv[15]))

            # CHECK DATA INFO INDEXES
            if not argv[16].isnumeric() or int(argv[16]) <= 0:
                self._load_data['index_errors'].append(('DEMOGRAPHIC_VARS_ROW', argv[16]))
            if not argv[17].isnumeric() or int(argv[17]) <= 0:
                self._load_data['index_errors'].append(('DEMOGRAPHIC_VARS_COL', argv[17]))

            # CHECK FILTER INDEXES
            if not argv[18].isnumeric() or int(argv[18]) <= 0:
                self._load_data['index_errors'].append(('FILTER_VARS_ROW', argv[18]))
            if not argv[19].isnumeric() or int(argv[19]) <= 0:
                self._load_data['index_errors'].append(('FILTER_VARS_COL', argv[19]))

            if not len(self._load_data['index_errors']) and self._load_data['size'] > 0:

                # READ EXCEL COLUMN
                def read_column(row_start, column_start, row_end=None):
                    res_column = []

                    # GET LAST ROW INDEX
                    if row_end is None:
                        row_index = row_start
                        if self._load_data['format'] == Loader._XLSX_FORMAT:
                            while True:
                                cell_value = self._sheet.cell(row_index, column_start).value
                                if cell_value is None or cell_value == '':
                                    row_end = row_index
                                    break
                                row_index += 1

                        elif self._load_data['format'] == Loader._XLS_FORMAT:
                            while True:
                                cell_value = self._sheet.cell_value(row_index - 1, column_start - 1)
                                if cell_value is None or cell_value == '':
                                    row_end = row_index
                                    break
                                row_index += 1
                    elif not isinstance(row_end, str):
                        row_end = row_start + row_end

                    # READ COLUMN
                    if isinstance(row_end, int):
                        if self._load_data['format'] == Loader._XLSX_FORMAT:
                            res_column = [self._sheet.cell(_, column_start).value for _ in range(row_start, row_end)]
                        elif self._load_data['format'] == Loader._XLS_FORMAT:
                            res_column = [self._sheet.cell_value(_, column_start - 1) for _ in range(row_start - 1, row_end - 1)]
                    elif isinstance(row_end, str):
                        row_index = row_start
                        if self._load_data['format'] == Loader._XLSX_FORMAT:
                            while True:
                                cell_value = self._sheet.cell(row_index, column_start).value
                                if cell_value == row_end:
                                    break
                                res_column.append(cell_value)
                                row_index += 1
                        elif self._load_data['format'] == Loader._XLS_FORMAT:
                            while True:
                                cell_value = self._sheet.cell_value(row_index - 1, column_start - 1)
                                if cell_value == row_end:
                                    break
                                res_column.append(cell_value)
                                row_index += 1

                    return res_column

                # READ ALERT FORM DATA
                self._load_data['brand_list'] = \
                    [''.join(_ for _ in cell_value if _.isalnum() or _ == ' ') for cell_value in read_column(int(argv[6]), int(argv[7]))]
                self._load_data['num_brand_list'] = read_column(int(argv[6]), int(argv[7]) - 1)

                self._load_data['effects_list'] = \
                    [''.join(_ for _ in cell_value if _.isalnum() or _ == ' ') for cell_value in read_column(int(argv[8]), int(argv[9]))]
                self._load_data['num_effects_list'] = read_column(int(argv[8]), int(argv[9]) - 1)

                self._load_data['barriers_list'] = \
                    [''.join(_ for _ in cell_value if _.isalnum() or _ == ' ') for cell_value in read_column(int(argv[10]), int(argv[11]))]
                self._load_data['num_barriers_list'] = read_column(int(argv[10]), int(argv[11]) - 1)

                self._load_data['imagery_list'] = \
                    [''.join(_ for _ in cell_value if _.isalnum() or _ == ' ') for cell_value in read_column(int(argv[12]), int(argv[13]))]
                self._load_data['num_imagery_list'] = read_column(int(argv[12]), int(argv[13]) - 1)

                self._load_data['data_info'] = []
                for _ in read_column(int(argv[14]), int(argv[15]), row_end=9):
                    if _ is not None:
                        self._load_data['data_info'].append(_.replace(' ', ''))
                    else:
                        self._load_data['data_info'].append(_)


                self._load_data['demo_vars'] = read_column(int(argv[16]), int(argv[17]), row_end='SIM Filter variables')
                self._load_data['demo_vars'] = [_.replace(' ', '') for _ in self._load_data['demo_vars'] if _ is not None]

                self._load_data['filter_vars'] = read_column(int(argv[18]), int(argv[19]))
                self._load_data['filter_vars'] = [_.replace(' ', '') for _ in self._load_data['filter_vars'] if _ is not None]


class Output:
    __ERROR_REMAP_FILE = 'Wrong Remap path'
    __ERROR_REMAP_FORMAT = 'Wrong Remap file format'
    __ERROR_REMAP_EMPTY = 'Empty Remap'
    __ERROR_REMAP_HEADER = 'Incorrect Remap header'

    __ERROR_ALERT_FORM_FILE = 'Wrong Alert Form path'
    __ERROR_ALERT_FORM_FORMAT = 'Wrong Alert Form file format'
    __ERROR_ALERT_FORM_EMPTY = 'Empty Alert Form'

    __ERROR_OUTPUT_FILE = 'Wrong Output path'

    __ERROR_EMPTY_BRAND_LIST = 'Empty Brand list'
    __ERROR_EMPTY_EFFECTS_LIST = 'Empty Market Effects list'

    __ERROR_EMPTY_NUM_BRAND_LIST = 'Empty numeration in Brand list'
    __ERROR_EMPTY_NUM_EFFECTS_LIST = 'Empty numeration in Market Effects list'
    __ERROR_EMPTY_NUM_BARRIERS_LIST = 'Empty numeration in Barriers to Consideration list'
    __ERROR_EMPTY_NUM_IMAGERY_LIST = 'Empty numeration in Imagery list'

    __ERROR_WRONG_NUM_BRAND_LIST = 'Wrong numeration in Brand list'
    __ERROR_WRONG_NUM_EFFECTS_LIST = 'Wrong numeration in Market Effects list'
    __ERROR_WRONG_NUM_BARRIERS_LIST = 'Wrong numeration in Barriers to Consideration list'
    __ERROR_WRONG_NUM_IMAGERY_LIST = 'Wrong numeration in Imagery list'

    __ERROR_LARGE_BRAND_LIST = 'String limit 50 chars in Brand list'
    __ERROR_LARGE_EFFECTS_LIST = 'String limit 50 chars in Market Effects list'
    __ERROR_LARGE_BARRIERS_LIST = 'String limit 50 chars in Barriers to Consideration list'
    __ERROR_LARGE_IMAGERY_LIST = 'String limit 50 chars in Imagery list'

    __ERROR_VARS_LABELS = 'Empty Value labels'
    __ERROR_VARS_NOT_FOUND = 'Variable is not found'
    __ERROR_VARS_SIZE = 'Not enough variables in'

    __ERROR_ARGUMENT = 'Wrong argument'
    __ERROR_FORMAT = 'Wrong Data Information format in'

    __ERROR_EMPTY_DATA_VAR = 'Empty Data information'

    __WARNING_WRONG_WEIGHT = 'WEIGHT is not found'
    __WARNING_WRONG_WAVE = 'WAVE is not found'

    __WARNING_EMPTY_DATA_VAR = 'Empty Data information'
    __WARNING_EMPTY_DEM_VARS = 'Empty SIM Demographic variables list'
    __WARNING_EMPTY_DATA_VAR = 'Empty Data information'

    __WARNING_EMPTY_BARRIERS_LIST = 'Empty Barriers to Consideration list'
    __WARNING_EMPTY_IMAGERY_LIST = 'Empty Imagery list'

    __WINDOW_WIDTH = 1000
    __WINDOW_HEIGHT = 400

    def __init__(self, path, rm_data, af_data):
        self.__rm_data = rm_data
        self.__af_data = af_data
        self.__num_line = 1

        # CREATE EMPTY DATA
        self.__load_data = defaultdict(lambda: None)

        # CREATE WINDOW
        self.__window = tkinter.Tk()
        self.__window.title('')
        right = self.__window.winfo_screenwidth() / 2 - self.__window.winfo_reqwidth() / 2
        down = self.__window.winfo_screenheight() / 2 - self.__window.winfo_reqheight() / 2
        self.__window.geometry(
            "{}x{}+{}+{}".format(Output.__WINDOW_WIDTH, Output.__WINDOW_HEIGHT,
                                 int(right - Output.__WINDOW_WIDTH / 2),
                                 int(down - Output.__WINDOW_HEIGHT / 2)))
        self.__text_area = tkinter.Text(font=("Consolas", 12))
        scroll_bar = tkinter.Scrollbar(self.__window)
        scroll_bar.config(command=self.__text_area.yview)
        self.__text_area.config(yscrollcommand=scroll_bar.set)
        scroll_bar.pack(side=tkinter.RIGHT, fill=tkinter.Y)
        self.__text_area.pack(expand=tkinter.YES, fill=tkinter.BOTH)
        self.__text_area.configure(background="black")

        # CHECK OUTPUT PATH
        self.__load_data['is_valid_file'] = path.endswith('.xlsx')
        if self.__load_data['is_valid_file']:
            try:
                openpyxl.Workbook().save(path)
            except OSError:
                self.__load_data['is_valid_file'] = False

            if self.__load_data['is_valid_file']:
                remove(path)

                self.__wb_output = openpyxl.Workbook()
                self.__ws_output = self.__wb_output.active
                self.__output_row = 6

                # CREATE OUTPUT FONTS
                self.__font1 = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE)
                self.__font2 = openpyxl.styles.Font(bold=True)
                self.__font3 = openpyxl.styles.Font(color=openpyxl.styles.colors.BLUE, bold=True)

                # LOAD REMAP VARIABLES
                try:
                    self.__vars = sorted(rm_data['vars'], key=lambda x: x[0])
                    self.__len_vars = len(self.__vars)
                except TypeError:
                    self.__vars = []
                    self.__len_vars = 0

                # CHECK WEIGHT NAME
                self.__load_data['is_weight'] = bool(len(argv[4]))
                if self.__load_data['is_weight']:
                    self.__load_data['is_valid_weight'] = \
                        self._binary_search(0, self.__len_vars - 1, argv[4]) is not None

                # CHECK WAVE NAME
                self.__load_data['is_wave'] = bool(len(argv[5]))
                if self.__load_data['is_wave']:
                    self.__load_data['is_valid_wave'] = \
                        self._binary_search(0, self.__len_vars - 1, argv[5]) is not None

                # GET FORMATS IN DATA INFORMATION
                def get_formated_name(var_name):
                    if var_name is None:
                        return None
                    if var_name.count('[') == 2 and var_name.count(']') == 2:
                        index_opens = [_ for _, v in enumerate(var_name) if v == '[']
                        index_closes = [_ for _, v in enumerate(var_name) if v == ']']
                        indexes = [index_opens[0], index_closes[0], index_opens[1], index_closes[1]]
                        if not ((index_closes[0] + 1 != index_opens[1]) or (index_closes[1] != len(var_name) - 1) or (indexes != sorted(indexes))):
                            formats = re.findall(r'\[.*?\]', var_name)
                            return var_name[:var_name.index('[')], formats[0][1:len(formats[0])-1], formats[1][1:len(formats[1])-1]
                    elif var_name.count('[') == 0 and var_name.count(']') == 0:
                        return var_name, '_', ''

                if self.__af_data['is_valid_file'] and (self.__af_data['format'] == Loader._XLSX_FORMAT
                                                        or self.__af_data['format'] == Loader._XLS_FORMAT) and not len(
                    self.__af_data['index_errors']) and self.__af_data['size'] > 0:

                    self.__load_data['total_awareness'] = \
                        self._get_interval(self.__af_data['data_info'][0]) if self.__af_data['data_info'][0] is not None else []

                    self.__load_data['usage'] = \
                        self._get_interval(self.__af_data['data_info'][1]) if self.__af_data['data_info'][1] is not None else []

                    self.__af_data['sow_formated_name'] = get_formated_name(self.__af_data['data_info'][2])
                    self.__load_data['share_of_wallet'] = \
                        self._get_numerical_interval(self.__af_data['sow_formated_name']) if self.__af_data['sow_formated_name'] is not None else []

                    self.__load_data['consideration'] = \
                        self._get_interval(self.__af_data['data_info'][3]) if self.__af_data['data_info'][3] is not None else []

                    self.__af_data['bp_formated_name'] = get_formated_name(self.__af_data['data_info'][4])
                    self.__load_data['brand_performance'] = \
                        self._get_numerical_interval(self.__af_data['bp_formated_name']) if self.__af_data['bp_formated_name'] is not None else []

                    self.__af_data['c_formated_name'] = get_formated_name(self.__af_data['data_info'][5])
                    self.__load_data['closeness'] = \
                        self._get_numerical_interval(self.__af_data['c_formated_name']) if self.__af_data['c_formated_name'] is not None else []

                    self.__af_data['me_formated_name'] = get_formated_name(self.__af_data['data_info'][6])
                    self.__load_data['market_effects'] =\
                        self._get_multi_interval(self.__af_data['me_formated_name'],
                                                 len(self.__af_data['effects_list'])) if self.__af_data['me_formated_name'] is not None else []

                    self.__af_data['btc_formated_name'] = get_formated_name(self.__af_data['data_info'][7])
                    self.__load_data['barriers_to_consideration'] = \
                        self._get_multi_interval(self.__af_data['btc_formated_name'],
                                                 len(self.__af_data['barriers_list'])) if self.__af_data['btc_formated_name'] is not None else []

                    self.__af_data['i_formated_name'] = get_formated_name(self.__af_data['data_info'][8])
                    self.__load_data['imagery'] = \
                        self._get_multi_interval(self.__af_data['i_formated_name'],
                                                 len(self.__af_data['imagery_list'])) if self.__af_data['i_formated_name'] is not None else []

                    self.__load_data['demo_vars'] = []
                    for _ in self.__af_data['demo_vars']:
                        self.__load_data['demo_vars'].append(self._get_variable_interval(_))

                    self.__load_data['filter_vars'] = []
                    for _ in self.__af_data['filter_vars']:
                        self.__load_data['filter_vars'].append(self._get_variable_interval(_))

        is_errors, is_warnings = self._get_errors()

        if is_errors or is_warnings:
            self._show_output()

        if not is_errors:

            # COPY EXCEL ROW
            def copy_row(num_row, start, end):
                if self.__rm_data['format'] == 'xlsx':
                    for _ in range(start, end + 1):
                        cell_value = self.__rm_data['sheet'].cell(num_row, _).value
                        if cell_value is not None:
                            self.__ws_output.cell(row=self.__output_row, column=_).value = cell_value
                    self.__output_row += 1
                elif self.__rm_data['format'] == 'xls':
                    for _ in range(start, end + 1):
                        cell_value = self.__rm_data['sheet'].cell_value(num_row - 1, _ - 1)
                        if cell_value is not None:
                            self.__ws_output.cell(row=self.__output_row, column=_).value = cell_value
                    self.__output_row += 1

            # WRITE DATA INFORMATION TO REMAP
            def write_data_info(var_names, var_labels, command=True):
                for _ in range(len(var_names)):
                    self.__ws_output.cell(row=self.__output_row + _, column=1).value = var_names[_]
                    self.__ws_output.cell(row=self.__output_row + _, column=3).value = var_labels[_]
                    self.__ws_output.cell(row=self.__output_row + _, column=4).value = 2
                    self.__ws_output.cell(row=self.__output_row + _, column=5).value = 'N'
                    self.__ws_output.cell(row=self.__output_row + _, column=6).value = 1
                    if command:
                        self.__ws_output.cell(row=self.__output_row, column=9).value = 'recode ' + var_names[0] + ' to ' + var_names[-1] + ' (1 thru hi = 1)(else = 0).@'

            # ADD YES/NO ALTERNATIVES
            def add_yes_no():
                self.__ws_output.cell(row=self.__output_row, column=2).value = 0
                self.__ws_output.cell(row=self.__output_row, column=3).value = 'No'
                self.__ws_output.cell(row=self.__output_row + 1, column=2).value = 1
                self.__ws_output.cell(row=self.__output_row + 1, column=3).value = 'Yes'
                self.__output_row += 2

            # COPY HEADER
            copy_row(6, 1, 17)

            # RESPONDEND ID
            for i, v in enumerate(['SERIAL', '', 'Serial', '', 'N', 8, '', 'compute', 'compute serial=$casenum.@']):
                self.__ws_output.cell(row=self.__output_row, column=i + 1).value = v
            self.__ws_output.cell(row=self.__output_row, column=11).value = self.__rm_data['vars'][0][0]
            self.__output_row += 1

            # WEIGHT
            if self.__load_data['is_valid_weight']:
                interval = self._get_variable_interval(argv[4])
                for _ in range(interval[1], interval[2] + 1):
                    copy_row(_, 1, 17)
                self.__ws_output.cell(row=self.__output_row - (interval[2] - interval[1]) - 1, column=1).value = 'WEIGHT'
                self.__ws_output.cell(row=self.__output_row - (interval[2] - interval[1]) - 1, column=3).value = 'Weight'
                self.__ws_output.cell(row=self.__output_row - (interval[2] - interval[1]) - 1, column=6).value = 20
                self.__ws_output.cell(row=self.__output_row - (interval[2] - interval[1]) - 1, column=7).value = 15
            if not self.__load_data['is_weight'] or not self.__load_data['is_valid_weight']:
                for i, v in enumerate(['WEIGHT', '', 'Weight', '', 'N', 20, 15, 'compute', 'compute WEIGHT=1.@']):
                    self.__ws_output.cell(row=self.__output_row, column=1 + i).value = v
                self.__output_row += 1

            # WAVE
            if self.__load_data['is_valid_wave']:
                interval = self._get_variable_interval(argv[5])
                for _ in range(interval[1], interval[2] + 1):
                    copy_row(_, 1, 17)
                self.__ws_output.cell(row=self.__output_row - (interval[2] - interval[1]) - 1, column=1).value = 'WAVE'
                self.__ws_output.cell(row=self.__output_row - (interval[2] - interval[1]) - 1, column=3).value = 'Wave'
            else:
                for i, v in enumerate(['WAVE', '', 'Wave', 1, 'N', 1, '', 'compute', 'compute WAVE=1.@']):
                    self.__ws_output.cell(row=self.__output_row, column=1 + i).value = v
                self.__ws_output.cell(row=self.__output_row + 1, column=2).value = 1
                self.__ws_output.cell(row=self.__output_row + 1, column=3).value = 'Wave 1'
                self.__output_row += 2

            # TOTAL AWARENESS
            write_data_info(['AWARE_BRAND_' + str(_) for _ in range(1, len(self.__af_data['brand_list']) + 1)],
                            ['Awareness_' + _ for _ in self.__af_data['brand_list']])
            for _ in self.__load_data['total_awareness'][:len(self.__af_data['brand_list'])]:
                copy_row(_[1], 11, 17)

            # USAGE
            write_data_info(['USAGE_BRAND_' + str(_) for _ in range(1, len(self.__af_data['brand_list']) + 1)],
                            ['Usage_' + _ for _ in self.__af_data['brand_list']])
            for _ in self.__load_data['usage'][:len(self.__af_data['brand_list'])]:
                copy_row(_[1], 11, 17)

            # USAGE
            write_data_info(['CONSIDER_BRAND_' + str(_) for _ in range(1, len(self.__af_data['brand_list']) + 1)],
                            ['Would consider_' + _ for _ in self.__af_data['brand_list']])
            for _ in self.__load_data['consideration'][:len(self.__af_data['brand_list'])]:
                copy_row(_[1], 11, 17)

            add_yes_no()

            # BRAND PERFORMANCE
            write_data_info(['BVC_PERF_' + str(_) for _ in range(1, len(self.__af_data['brand_list']) + 1)],
                            ['Brand Performance_' + _ for _ in self.__af_data['brand_list']], command=False)
            for _ in self.__load_data['brand_performance'][:len(self.__af_data['brand_list'])]:
                copy_row(_[1], 6, 17)

            # CLAIMED SHARE
            write_data_info(['BVC_SOW_' + str(_) for _ in range(1, len(self.__af_data['brand_list']) + 1)],
                            ['Claimed Share_' + _ for _ in self.__af_data['brand_list']], command=False)
            for _ in self.__load_data['share_of_wallet'][:len(self.__af_data['brand_list'])]:
                copy_row(_[1], 6, 17)

            sow_decimals = []
            if self.__rm_data['format'] == 'xlsx':
                for _ in self.__load_data['share_of_wallet'][:len(self.__af_data['brand_list'])]:
                    sow_decimals.append(self.__rm_data['sheet'].cell(_[1], 17).value)
            elif self.__rm_data['format'] == 'xls':
                for _ in self.__load_data['share_of_wallet'][:len(self.__af_data['brand_list'])]:
                    sow_decimals.append(self.__rm_data['sheet'].cell_value(_[1] - 1, 16))
            if not all(sow_decimals):
                for _ in range(1, len(self.__load_data['share_of_wallet'][:len(self.__af_data['brand_list'])]) + 1):
                    self.__ws_output.cell(row=self.__output_row - _, column=6).value = 6
                    self.__ws_output.cell(row=self.__output_row - _, column=7).value = 2

            # CLOSENESS
            write_data_info(['BVC_INV_' + str(_) for _ in range(1, len(self.__af_data['brand_list']) + 1)],
                            ['Closeness_' + _ for _ in self.__af_data['brand_list']], command=False)
            for _ in self.__load_data['closeness'][:len(self.__af_data['brand_list'])]:
                copy_row(_[1], 6, 17)

            # MARKET EFFECTS
            var_names, var_labels = [], []
            for mei, mev in enumerate(self.__af_data['effects_list']):
                for bli, blv in enumerate(self.__af_data['brand_list']):
                    var_names.append('BAR_' + str(mei + 1) + '_BRAND_' + str(bli + 1))
                    var_labels.append(mev + '_' + blv)
            write_data_info(var_names, var_labels)
            for i in self.__load_data['market_effects']:
                for j in i[:len(self.__af_data['brand_list'])]:
                    copy_row(j[1], 11, 17)

            add_yes_no()

            # DEMOGRAPHIC VARIABLES
            for i in self.__load_data['demo_vars']:
                self.__ws_output.cell(row=self.__output_row, column=1).value = 'DEM_' + i[0]
                for j in range(i[1], i[2]+1):
                    copy_row(j, 2, 17)

            # FILTER VARIABLES
            for i in self.__load_data['filter_vars']:
                self.__ws_output.cell(row=self.__output_row, column=1).value = 'FIL_' + i[0]
                for j in range(i[1], i[2] + 1):
                    copy_row(j, 2, 17)

            # IMAGERY
            is_imagery = self.__af_data['data_info'][8] is not None and len(self.__af_data['data_info'][8]) and len(self.__af_data['imagery_list'])
            if is_imagery:
                var_names, var_labels = [], []
                for mei, mev in enumerate(self.__af_data['imagery_list']):
                    for bli, blv in enumerate(self.__af_data['brand_list']):
                        var_names.append('ATT_' + str(mei + 1) + '_BRAND_' + str(bli + 1))
                        var_labels.append(mev + '_' + blv)
                write_data_info(var_names, var_labels)
                for i in self.__load_data['imagery']:
                    for j in i[:len(self.__af_data['brand_list'])]:
                        copy_row(j[1], 11, 17)

            # BARRIERS TO CONSIDERATION
            is_bar_consid = self.__af_data['data_info'][7] is not None and len(self.__af_data['data_info'][7]) and len(self.__af_data['barriers_list'])
            if is_bar_consid:
                var_names, var_labels = [], []
                for mei, mev in enumerate(self.__af_data['barriers_list']):
                    for bli, blv in enumerate(self.__af_data['brand_list']):
                        var_names.append('BARCON_' + str(mei + 1) + '_BRAND_' + str(bli + 1))
                        var_labels.append(mev + '_' + blv)
                write_data_info(var_names, var_labels)
                for i in self.__load_data['barriers_to_consideration']:
                    for j in i[:len(self.__af_data['brand_list'])]:
                        copy_row(j[1], 11, 17)

            if is_imagery or is_bar_consid:
                add_yes_no()

            # FORMAT REMAP VIEW
            for _ in range(7, self.__ws_output.max_row + 1):
                self.__ws_output.cell(row=_, column=1).font = self.__font1
                self.__ws_output.cell(row=_, column=11).font = self.__font1
                self.__ws_output.cell(row=_, column=8).font = self.__font2
                self.__ws_output.cell(row=_, column=9).font = self.__font2
            for _ in range(1, 5):
                self.__ws_output.cell(row=_, column=11).font = self.__font2
            for _ in range(1, 18):
                self.__ws_output.cell(row=6, column=_).font = self.__font2
            self.__ws_output.cell(row=6, column=1).font = self.__font3
            self.__ws_output.cell(row=6, column=11).font = self.__font3
            self.__ws_output.column_dimensions['A'].width = 22
            self.__ws_output.column_dimensions['C'].width = 50
            self.__ws_output.column_dimensions['H'].width = 12
            self.__ws_output.column_dimensions['I'].width = 30
            self.__ws_output.column_dimensions['K'].width = 20
            self.__ws_output.column_dimensions['M'].width = 50
            self.__ws_output.freeze_panes = self.__ws_output['B7']

            # SAVE FILE
            self.__wb_output.save(argv[3])


    def _add_line(self, text, color):
        line_num_str = str(self.__num_line)
        self.__text_area.insert(tkinter.INSERT, str(self.__num_line) + ': ' + text + '\n')
        self.__text_area.tag_add(line_num_str, line_num_str + '.0', line_num_str + '.end')
        self.__text_area.tag_config(line_num_str, foreground=color)
        self.__num_line += 1

    def _get_errors(self):
        errors = []
        warnings = []

        # CHECK FILES AND FORMATS
        if not self.__rm_data['is_valid_file'] or (
        not (self.__rm_data['format'] == 'xls' or self.__rm_data['format'] == 'xlsx')):
            errors.append(Output.__ERROR_REMAP_FILE)
        if not self.__af_data['is_valid_file'] or (
        not (self.__af_data['format'] == 'xls' or self.__af_data['format'] == 'xlsx')):
            errors.append(Output.__ERROR_ALERT_FORM_FILE)
        if not self.__load_data['is_valid_file']:
            errors.append(Output.__ERROR_OUTPUT_FILE)

        # CHECK BAT-FILE
        if self.__af_data['index_errors'] is not None:
            for _ in self.__af_data['index_errors']:
                errors.append(Output.__ERROR_ARGUMENT + ' - ' + _[0] + ' = ' + _[1])

        if not len(errors):

            # CHECK EMPTY REMAP AND HEADER
            if self.__rm_data['is_empty']:
                errors.append(Output.__ERROR_REMAP_EMPTY)
            elif not self.__rm_data['is_valid_header']:
                errors.append(Output.__ERROR_REMAP_HEADER)

            # CHECK EMPTY ALERT FORM
            if self.__af_data['is_empty']:
                errors.append(Output.__ERROR_ALERT_FORM_EMPTY)
            elif not self.__rm_data['is_empty'] and self.__rm_data['is_valid_header']:
                # CHECK BRAND LIST
                if not len(self.__af_data['brand_list']):
                    errors.append(Output.__ERROR_EMPTY_BRAND_LIST)
                else:
                    for _ in self.__af_data['brand_list']:
                        if len(_) > 50:
                            errors.append(Output.__ERROR_LARGE_BRAND_LIST + ' - ' + _)

                # CHECK MARKET EFFECTS LIST
                if not len(self.__af_data['effects_list']):
                    errors.append(Output.__ERROR_EMPTY_EFFECTS_LIST)
                else:
                    for _ in self.__af_data['effects_list']:
                        if len(_) > 50:
                            errors.append(Output.__ERROR_LARGE_EFFECTS_LIST + ' - ' + _)

                # CHECK BARRIERS LIST
                if len(self.__af_data['barriers_list']):
                    for _ in self.__af_data['barriers_list']:
                        if len(_) > 50:
                            errors.append(Output.__ERROR_LARGE_BARRIERS_LIST + ' - ' + _)

                # CHECK IMAGERY LIST
                if len(self.__af_data['imagery_list']):
                    for _ in self.__af_data['imagery_list']:
                        if len(_) > 50:
                            errors.append(Output.__ERROR_LARGE_IMAGERY_LIST + ' - ' + _)

                # CHECK BRANDS NUMERATION
                if len(self.__af_data['brand_list']):
                    if not len(self.__af_data['num_brand_list']):
                        errors.append(Output.__ERROR_WRONG_NUM_BRAND_LIST)
                    elif len(self.__af_data['brand_list']) != len(self.__af_data['num_brand_list']):
                        errors.append(Output.__ERROR_WRONG_NUM_BRAND_LIST)
                    else:
                        try:
                            for num, value in enumerate(self.__af_data['num_brand_list']):
                                if num + 1 != int(value):
                                    errors.append(Output.__ERROR_WRONG_NUM_BRAND_LIST)
                                    break
                        except ValueError:
                            errors.append(Output.__ERROR_WRONG_NUM_BRAND_LIST)

                # CHECK MARKET EFFECTS NUMERATION
                if len(self.__af_data['effects_list']):
                    if not len(self.__af_data['num_effects_list']):
                        errors.append(Output.__ERROR_WRONG_NUM_EFFECTS_LIST)
                    elif len(self.__af_data['effects_list']) != len(self.__af_data['num_effects_list']):
                        errors.append(Output.__ERROR_WRONG_NUM_EFFECTS_LIST)
                    else:
                        try:
                            for num, value in enumerate(self.__af_data['num_effects_list']):
                                if num + 1 != int(value):
                                    errors.append(Output.__ERROR_WRONG_NUM_EFFECTS_LIST)
                                    break
                        except ValueError:
                            errors.append(Output.__ERROR_WRONG_NUM_EFFECTS_LIST)

                # CHECK BARRIERS NUMERATION
                if len(self.__af_data['barriers_list']):
                    if not len(self.__af_data['num_barriers_list']):
                        errors.append(Output.__ERROR_WRONG_NUM_BARRIERS_LIST)
                    elif len(self.__af_data['barriers_list']) != len(self.__af_data['num_barriers_list']):
                        errors.append(Output.__ERROR_WRONG_NUM_BARRIERS_LIST)
                    else:
                        try:
                            for num, value in enumerate(self.__af_data['num_barriers_list']):
                                if num + 1 != int(value):
                                    errors.append(Output.__ERROR_WRONG_NUM_BARRIERS_LIST)
                                    break
                        except ValueError:
                            errors.append(Output.__ERROR_WRONG_NUM_BARRIERS_LIST)

                # CHECK IMAGERY NUMERATION
                if len(self.__af_data['imagery_list']):
                    if not len(self.__af_data['num_imagery_list']):
                        errors.append(Output.__ERROR_WRONG_NUM_IMAGERY_LIST)
                    elif len(self.__af_data['imagery_list']) != len(self.__af_data['num_imagery_list']):
                        errors.append(Output.__ERROR_WRONG_NUM_IMAGERY_LIST)
                    else:
                        try:
                            for num, value in enumerate(self.__af_data['num_imagery_list']):
                                if num + 1 != int(value):
                                    errors.append(Output.__ERROR_WRONG_NUM_IMAGERY_LIST)
                                    break
                        except ValueError:
                            errors.append(Output.__ERROR_WRONG_NUM_IMAGERY_LIST)

                # CHECK DATA INFORMATION VARS AND FORMATS AND SIZES
                if len(self.__af_data['brand_list']):
                    if self.__af_data['data_info'][0] is None or self.__af_data['data_info'][0] == '':
                        errors.append(Output.__ERROR_EMPTY_DATA_VAR + ' - ' + 'Total Awareness')
                    elif not len(self.__load_data['total_awareness']):
                        errors.append(
                            Output.__ERROR_VARS_NOT_FOUND + ' - Total Awareness - ' + self.__af_data['data_info'][0])
                    elif len(self.__load_data['total_awareness']) < len(self.__af_data['brand_list']):
                        errors.append(Output.__ERROR_VARS_SIZE + ' Total Awareness - ' + self.__af_data['data_info'][0])

                    if self.__af_data['data_info'][1] is None or self.__af_data['data_info'][1] == '':
                        errors.append(Output.__ERROR_EMPTY_DATA_VAR + ' - ' + 'Usage')
                    elif not len(self.__load_data['usage']):
                        errors.append(Output.__ERROR_VARS_NOT_FOUND + ' - Usage - ' + self.__af_data['data_info'][1])
                    elif len(self.__load_data['usage']) < len(self.__af_data['brand_list']):
                        errors.append(Output.__ERROR_VARS_SIZE + ' Usage - ' + self.__af_data['data_info'][1])

                    if self.__af_data['data_info'][2] is None or self.__af_data['data_info'][2] == '':
                        errors.append(Output.__ERROR_EMPTY_DATA_VAR + ' - ' + 'Share of Wallet')
                    elif self.__af_data['sow_formated_name'] is None:
                        errors.append(Output.__ERROR_FORMAT + ' Share of Wallet - ' + self.__af_data['data_info'][2])
                    elif not len(self.__load_data['share_of_wallet']):
                        errors.append(
                            Output.__ERROR_VARS_NOT_FOUND + ' - Share of Wallet - ' + self.__af_data['data_info'][2])
                    elif len(self.__load_data['share_of_wallet']) < len(self.__af_data['brand_list']):
                        errors.append(Output.__ERROR_VARS_SIZE + ' Share of Wallet - ' + self.__af_data['data_info'][2])

                    if self.__af_data['data_info'][3] is None or self.__af_data['data_info'][3] == '':
                        errors.append(Output.__ERROR_EMPTY_DATA_VAR + ' - ' + 'Consideration')
                    elif not len(self.__load_data['consideration']):
                        errors.append(
                            Output.__ERROR_VARS_NOT_FOUND + ' - Consideration - ' + self.__af_data['data_info'][3])
                    elif len(self.__load_data['consideration']) < len(self.__af_data['brand_list']):
                        errors.append(Output.__ERROR_VARS_SIZE + ' Consideration - ' + self.__af_data['data_info'][3])

                    if self.__af_data['data_info'][4] is None or self.__af_data['data_info'][4] == '':
                        errors.append(Output.__ERROR_EMPTY_DATA_VAR + ' - ' + 'Brand Performance')
                    elif self.__af_data['bp_formated_name'] is None:
                        errors.append(Output.__ERROR_FORMAT + ' Brand Performance - ' + self.__af_data['data_info'][4])
                    elif not len(self.__load_data['brand_performance']):
                        errors.append(
                            Output.__ERROR_VARS_NOT_FOUND + ' - Brand Performance - ' + self.__af_data['data_info'][4])
                    elif len(self.__load_data['brand_performance']) < len(self.__af_data['brand_list']):
                        errors.append(Output.__ERROR_VARS_SIZE + ' Brand Performance - ' + self.__af_data['data_info'][4])

                    if self.__af_data['data_info'][5] is None or self.__af_data['data_info'][5] == '':
                        errors.append(Output.__ERROR_EMPTY_DATA_VAR + ' - ' + 'Closeness')
                    elif self.__af_data['c_formated_name'] is None:
                        errors.append(Output.__ERROR_FORMAT + ' Closeness - ' + self.__af_data['data_info'][5])
                    elif not len(self.__load_data['closeness']):
                        errors.append(Output.__ERROR_VARS_NOT_FOUND + ' - Closeness - ' + self.__af_data['data_info'][5])
                    elif len(self.__load_data['closeness']) < len(self.__af_data['brand_list']):
                        errors.append(Output.__ERROR_VARS_SIZE + ' Closeness - ' + self.__af_data['data_info'][5])

                if len(self.__af_data['effects_list']):
                    if self.__af_data['data_info'][6] is None or self.__af_data['data_info'][6] == '':
                        errors.append(Output.__ERROR_EMPTY_DATA_VAR + ' - ' + 'Market Effects')
                    elif self.__af_data['me_formated_name'] is None:
                        errors.append(Output.__ERROR_FORMAT + ' Market Effects - ' + self.__af_data['data_info'][6])
                    elif not any(self.__load_data['market_effects']):
                        errors.append(
                            Output.__ERROR_VARS_NOT_FOUND + ' - Market Effects - ' + self.__af_data['data_info'][6])
                    elif len(self.__load_data['market_effects']) < len(self.__af_data['effects_list']):
                        errors.append(Output.__ERROR_VARS_SIZE + ' Market Effects - ' + self.__af_data['data_info'][6])
                    else:
                        for _ in self.__load_data['market_effects']:
                            if len(_) < len(self.__af_data['brand_list']):
                                errors.append(Output.__ERROR_VARS_SIZE + ' Market Effects - ' + self.__af_data['data_info'][6])
                                break

                if self.__af_data['data_info'][7] is None or self.__af_data['data_info'][7] == '':
                    warnings.append(Output.__WARNING_EMPTY_DATA_VAR + ' - ' + 'Barriers to Consideration')
                elif self.__af_data['btc_formated_name'] is None:
                    errors.append(Output.__ERROR_FORMAT + ' Barriers to Consideration - ' + self.__af_data['data_info'][7])
                elif not any(self.__load_data['barriers_to_consideration']):
                    errors.append(
                        Output.__ERROR_VARS_NOT_FOUND + ' - Barriers to Consideration - ' + self.__af_data['data_info'][
                            7])
                elif len(self.__load_data['barriers_to_consideration']) < len(self.__af_data['barriers_list']):
                    errors.append(Output.__ERROR_VARS_SIZE + ' Barriers to Consideration - ' + self.__af_data['data_info'][7])
                else:
                    for _ in self.__load_data['barriers_to_consideration']:
                        if len(_) < len(self.__af_data['brand_list']):
                            errors.append(Output.__ERROR_VARS_SIZE + ' Barriers to Consideration - ' + self.__af_data['data_info'][7])
                            break

                if self.__af_data['data_info'][8] is None or self.__af_data['data_info'][8] == '':
                    warnings.append(Output.__WARNING_EMPTY_DATA_VAR + ' - ' + 'Imagery')
                elif self.__af_data['i_formated_name'] is None:
                    errors.append(Output.__ERROR_FORMAT + ' Imagery - ' + self.__af_data['data_info'][8])
                elif not any(self.__load_data['imagery']):
                    errors.append(Output.__ERROR_VARS_NOT_FOUND + ' - Imagery - ' + self.__af_data['data_info'][8])
                elif len(self.__load_data['imagery']) < len(self.__af_data['imagery_list']):
                    errors.append(Output.__ERROR_VARS_SIZE + ' Imagery - ' + self.__af_data['data_info'][8])
                else:
                    for _ in self.__load_data['imagery']:
                        if len(_) < len(self.__af_data['brand_list']):
                            errors.append(Output.__ERROR_VARS_SIZE + ' Imagery - ' + self.__af_data['data_info'][8])
                            break

                # CHECK DEMOGRAPHIC VARIABLES
                for i, v in enumerate(self.__load_data['demo_vars']):
                    if v is None:
                        errors.append(
                            Output.__ERROR_VARS_NOT_FOUND + ' - Demographic variables - ' + self.__af_data['demo_vars'][
                                i])
                    #elif v[1] == v[2]:
                    #    errors.append(Output.__ERROR_VARS_LABELS + ' - ' + v[0])

                # CHECK FILTER VARIABLES
                for i, v in enumerate(self.__load_data['filter_vars']):
                    if v is None:
                        errors.append(
                            Output.__ERROR_VARS_NOT_FOUND + ' - Filter variables - ' + self.__af_data['filter_vars'][i])
                    elif v[1] == v[2]:
                        errors.append(Output.__ERROR_VARS_LABELS + ' - ' + v[0])

                # CHECK BARRIERS LIST
                if not len(self.__af_data['barriers_list']):
                    warnings.append(Output.__WARNING_EMPTY_BARRIERS_LIST)

                # CHECK IMAGERY LIST
                if not len(self.__af_data['imagery_list']):
                    warnings.append(Output.__WARNING_EMPTY_IMAGERY_LIST)

                # CHECK WEIGHT
                if self.__load_data['is_weight'] and not self.__load_data['is_valid_weight']:
                    warnings.append(Output.__WARNING_WRONG_WEIGHT + ' - ' + argv[4])

                # CHECK WAVE
                if self.__load_data['is_wave'] and not self.__load_data['is_valid_wave']:
                    warnings.append(Output.__WARNING_WRONG_WAVE + ' - ' + argv[5])

        for _ in errors:
            self._add_line('Error: ' + _, 'red')

        for _ in warnings:
            self._add_line('Warning: ' + _, 'yellow')

        return bool(len(errors)), bool(len(warnings))

    def _show_output(self):
        self.__text_area.config(state=tkinter.DISABLED)
        self.__window.mainloop()

    def _binary_search(self, left, right, value):
        if right >= left:
            middle = (right + left) // 2
            if self.__vars[middle][0] == value:
                return self.__vars[middle][1]
            elif self.__vars[middle][0] > value:
                return self._binary_search(left, middle - 1, value)
            else:
                return self._binary_search(middle + 1, right, value)

    def _get_interval(self, var_name):
        res = []
    
        for _ in self.__vars:
            if _[0].startswith(var_name + 'c'):
                for _ in range(1, 1000):
                    name = var_name + 'c' + str(_)
                    index = self._binary_search(0, self.__len_vars - 1, name)
                    if index is not None:
                        res.append((name, index))
                break
        return res

    def _get_multi_interval(self, var_name, size):
        return [self._get_interval(var_name[0] + var_name[1] + str(_) + var_name[2]) for _ in range(1, size + 1)]

    def _get_numerical_interval(self, var_name):
        res = []
        if var_name is None:
            return res
        for _ in self.__vars:
            if _[0].startswith(var_name[0]):
                for _ in range(1, 1000):
                    name = var_name[0] + var_name[1] + str(_) + var_name[2]
                    index = self._binary_search(0, self.__len_vars - 1, name)
                    if index is not None:
                        res.append((name, index))
                break
        return res

    def _get_variable_interval(self, var_name):
        for _ in self.__vars:
            if _[0] == var_name:
                return _

    def get_data(self):
        return self.__load_data


if __name__ == '__main__':
    remap_path = argv[1].replace('\\\\', '\\')
    alert_form_path = argv[2].replace('\\\\', '\\')
    output_path = argv[3].replace('\\\\', '\\')

    if __DEBUG_BAT:
        print(argv)

        print('\n', 'REMAP PATH = ' + remap_path, 'ALERT FORM PATH = ' + alert_form_path,
              'OUTPUT PATH = ' + output_path, '\n',
              'WT = ' + argv[4], 'WAVE = ' + argv[5], '\n',
              'BRAND LIST ROW = ' + argv[6], 'BRAND LIST COL = ' + argv[7], '\n',
              'MARKET EFFECTS ROW = ' + argv[8], 'MARKET EFFECTS COL = ' + argv[9], '\n',
              'BARRIERS CONSIDERATION ROW = ' + argv[10], 'BARRIERS CONSIDERATION COL = ' + argv[11], '\n',
              'IMAGERY ROW = ' + argv[12], 'IMAGERY COL = ' + argv[13], '\n',
              'DATA INFORMATION_ROW = ' + argv[14], 'DATA INFORMATION COL = ' + argv[15], '\n',
              'DEMOGRAPHIC VARS ROW = ' + argv[16], 'DEMOGRAPHIC VARS COL = ' + argv[17], '\n',
              'FILTER VARS ROW = ' + argv[18], 'FILTER VARS COL = ' + argv[19], '\n', sep='\n')

    remap_loader = RemapLoader(remap_path)
    remap_data = remap_loader.get_data()

    alert_form_loader = AlertFormLoader(alert_form_path)
    alert_form_data = alert_form_loader.get_data()

    output = Output(output_path, remap_data, alert_form_data)

    if __DEBUG_REMAP:
        print('--REMAP DATA--')
        print('VALID FILE : ' + str(remap_data['is_valid_file']))
        if remap_data['is_valid_file']:
            print('FORMAT : ' + remap_data['format'])
            print('SIZE: ' + str(remap_data['size']))
            print('EMPTY : ' + str(remap_data['is_empty']))

            if remap_data['format'] == 'xlsx' or remap_data['format'] == 'xls':
                print('VALID HEADER : ' + str(remap_data['is_valid_header']))
                if remap_data['is_valid_header'] and not remap_data['is_empty']:
                    print('\nVARIABLES: ')
                    for _ in remap_data['vars']:
                        print(_[0] + ' = ' + str(_[1]))

        print('\n----------------------------------------------------\n')

    if __DEBUG_ALERT_FORM:
        print('--ALERT FORM DATA--')
        print('VALID FILE : ' + str(alert_form_data['is_valid_file']))
        if alert_form_data['is_valid_file']:
            print('FORMAT : ' + alert_form_data['format'])
            print('SIZE: ' + str(alert_form_data['size']))
            print('EMPTY : ' + str(alert_form_data['is_empty']))

            if len(alert_form_data['index_errors']):
                print('\nINDEX ERRORS:')
                for _ in alert_form_data['index_errors']:
                    print(_[0] + ' : ' + _[1])
            else:
                print('\nBRAND LIST : ' + str(alert_form_data['brand_list']))
                print('\nBRAND LIST NUMERATION : ' + str(alert_form_data['num_brand_list']))
                print('\nEFFECTS LIST : ' + str(alert_form_data['effects_list']))
                print('\nEFFECTS LIST NUMERATION : ' + str(alert_form_data['num_effects_list']))
                print('\nBARRIERS LIST : ' + str(alert_form_data['barriers_list']))
                print('\nBARRIERS LIST NUMERATION : ' + str(alert_form_data['num_barriers_list']))
                print('\nIMAGERY LIST : ' + str(alert_form_data['imagery_list']))
                print('\nIMAGERY LIST NUMERATION : ' + str(alert_form_data['num_imagery_list']))
                print('\nDATA INFORMATION : ' + str(alert_form_data['data_info']))
                print('\nDEMOGRAPHIC VARIABLES : ' + str(alert_form_data['demo_vars']))
                print('\nFILTER VARIABLES : ' + str(alert_form_data['filter_vars']))

        print('\n----------------------------------------------------\n')

    if __DEBUG_OUTPUT:
        output_data = output.get_data()

        print('--OUTPUT DATA--')
        print('VALID FILE : ' + str(output_data['is_valid_file']))
        print('SET WEIGHT = ' + str(output_data['is_weight']))
        if output_data['is_weight']:
            print('VALID WEIGHT = ' + str(output_data['is_valid_weight']))
        print('SET WAVE = ' + str(output_data['is_wave']))
        if output_data['is_wave']:
            print('VALID WAVE = ' + str(output_data['is_valid_wave']))

        if output_data['is_valid_file']:
            print('\nTOTAL AWARENESS ROWS : ')
            if output_data['total_awareness'] is not None:
                for _ in output_data['total_awareness']:
                    print(_[0] + ' : ' + str(_[1]))
            print('\nUSAGE ROWS : ')
            if output_data['usage'] is not None:
                for _ in output_data['usage']:
                    print(_[0] + ' : ' + str(_[1]))
            print('\nSHARE OF WALLET ROWS : ')
            if output_data['share_of_wallet'] is not None:
                for _ in output_data['share_of_wallet']:
                    print(_[0] + ' : ' + str(_[1]))
            print('\nCONSIDERATION ROWS : ')
            if output_data['consideration'] is not None:
                for _ in output_data['consideration']:
                    print(_[0] + ' : ' + str(_[1]))
            print('\nBRAND PERFORMANCE ROWS : ')
            if output_data['brand_performance'] is not None:
                for _ in output_data['brand_performance']:
                    print(_[0] + ' : ' + str(_[1]))
            print('\nCLOSENESS ROWS : ')
            if output_data['closeness'] is not None:
                for _ in output_data['closeness']:
                    print(_[0] + ' : ' + str(_[1]))
            print('\nMARKET EFFECTS ROWS : ')
            if output_data['market_effects'] is not None:
                for i in output_data['market_effects']:
                    for v in i:
                        print(v[0] + ' : ' + str(v[1]))
            print('\nBARRIERS TO CONSIDERATION ROWS : ')
            if output_data['barriers_to_consideration'] is not None:
                for i in output_data['barriers_to_consideration']:
                    for v in i:
                        print(v[0] + ' : ' + str(v[1]))
            print('\nIMAGERY ROWS : ')
            if output_data['imagery'] is not None:
                for i in output_data['imagery']:
                    for v in i:
                        print(v[0] + ' : ' + str(v[1]))
            print('\nDEMO VARS ROWS : ')
            if output_data['demo_vars'] is not None:
                for _ in output_data['demo_vars']:
                    if _ is not None:
                        print(_[0] + ' : ' + str(_[1]) + ' - ' + str(_[2]))
            print('\nFILTER VARS ROWS : ')
            if output_data['filter_vars'] is not None:
                for _ in output_data['filter_vars']:
                    if _ is not None:
                        print(_[0] + ' : ' + str(_[1]) + ' - ' + str(_[2]))
